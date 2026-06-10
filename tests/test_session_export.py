"""session_export 多 Sheet Excel 单元测试"""

import io

import pandas as pd
import pytest
from openpyxl import load_workbook

from utils.session_export import (
    build_export_dataframe,
    build_multi_sheet_excel_bytes,
    build_summary_kpi_dataframe,
    build_summary_ranking_dataframe,
)


def test_multi_sheet_excel_has_three_sheets():
    main = pd.DataFrame({"日期": ["2024-01-01"], "销售额": [100.0]})
    lookup = pd.DataFrame({"产品类型": ["电子产品"], "负责人": ["张明"]})
    metrics = {
        "total_sales": 100.0,
        "mom_growth_pct": 5.5,
        "order_count": 1,
        "warning_count": 0,
        "province_ranking": [{"name": "华东", "sales": 100.0}],
        "meta": {"date_col": "日期", "sales_col": "销售额", "region_col": "区域"},
    }
    kpi = build_summary_kpi_dataframe(
        filename="demo.xlsx",
        lookup_filename="lookup.xlsx",
        row_count=1,
        metrics=metrics,
    )
    ranking = build_summary_ranking_dataframe(metrics)

    payload = build_multi_sheet_excel_bytes(
        main,
        lookup_df=lookup,
        summary_kpi=kpi,
        summary_ranking=ranking,
    )
    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["数据", "参考表", "汇总"]


def test_multi_sheet_without_lookup():
    main = pd.DataFrame({"A": [1]})
    kpi = build_summary_kpi_dataframe(
        filename="t.csv",
        lookup_filename=None,
        row_count=1,
        metrics={"total_sales": 1, "order_count": 1, "warning_count": 0},
    )
    payload = build_multi_sheet_excel_bytes(main, summary_kpi=kpi)
    wb = load_workbook(io.BytesIO(payload))
    assert wb.sheetnames == ["数据", "汇总"]


def test_build_export_dataframe_respects_visible():
    df = pd.DataFrame({"日期": ["2024-01-01"], "销售额": [10], "隐藏列": [1]})
    meta = [
        {"name": "日期", "visible": True},
        {"name": "销售额", "visible": True},
        {"name": "隐藏列", "visible": False},
    ]
    out = build_export_dataframe(df, meta, "日期", "销售额", None)
    assert "隐藏列" not in out.columns
    assert len(out.columns) == 2
