"""
会话主表导出：按 mapped_columns 的显隐与顺序，并应用日期/数值/维度显示名。
不依赖 Flask app 模块，由路由传入 df 与 meta。
"""
from __future__ import annotations

import io
import logging
import os
import re
import sys
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

logger = logging.getLogger(__name__)

# 用户可将 SimHei.ttf / msyh.ttc 等放入此目录（相对项目根）
PDF_FONT_USER_DIR = "fonts"

PDF_FONT_HELP_ZH = (
    "未找到可用的中文字体，PDF 中的中文可能显示为方框。"
    f"请将 SimHei.ttf 或 msyh.ttc 复制到项目根下的「{PDF_FONT_USER_DIR}」文件夹后重试；"
    "Windows 用户也可在 C:\\Windows\\Fonts 安装黑体/微软雅黑。"
)

# ---- 构建待导出的 DataFrame（仅可见列 + 重命名表头） ----


def _display_header_for_column(
    col: str,
    item: Optional[Dict[str, Any]],
    md: Optional[str],
    mv: Optional[str],
    mdm: Optional[str],
) -> str:
    if item and isinstance(item.get("display_name"), str) and item["display_name"].strip():
        return item["display_name"].strip()
    if md and col == md:
        return "日期"
    if mv and col == mv:
        return "数值"
    if mdm and col == mdm:
        return "维度"
    return str(col)


def _dedupe_headers(pairs: List[Tuple[str, str]]) -> List[Tuple[str, str]]:
    """保证导出表头唯一，避免 pandas 重名列。"""
    used: Dict[str, int] = {}
    out: List[Tuple[str, str]] = []
    for orig, disp in pairs:
        d = disp
        if d in used:
            used[d] += 1
            d = f"{disp} ({used[d]})"
        else:
            used[d] = 1
        out.append((orig, d))
    return out


def build_export_dataframe(
    df_full: pd.DataFrame,
    layout_meta: List[Dict[str, Any]],
    mapped_date_col: Optional[str],
    mapped_value_col: Optional[str],
    mapped_dim_col: Optional[str],
) -> pd.DataFrame:
    """
    从全量 df 按 layout_meta 顺序仅取 visible 列，并按映射生成中文友好表头。
    layout_meta 建议为 normalize 后的列表（与 GLOBAL_DATA['mapped_columns'] 一致）。
    """
    if df_full is None or df_full.empty:
        raise ValueError("没有可导出的数据")

    by_name: Dict[str, Dict[str, Any]] = {}
    for item in layout_meta:
        if isinstance(item, dict) and item.get("name") in df_full.columns:
            by_name[str(item["name"])] = item

    pairs: List[Tuple[str, str]] = []
    seen: set = set()
    for item in layout_meta:
        if not isinstance(item, dict):
            continue
        name = item.get("name")
        if name is None or name not in df_full.columns or name in seen:
            continue
        seen.add(name)
        if not item.get("visible", True):
            continue
        disp = _display_header_for_column(
            str(name), item, mapped_date_col, mapped_value_col, mapped_dim_col
        )
        pairs.append((str(name), disp))

    for c in df_full.columns:
        if c not in seen:
            item = by_name.get(str(c))
            disp = _display_header_for_column(
                str(c), item, mapped_date_col, mapped_value_col, mapped_dim_col
            )
            pairs.append((str(c), disp))

    if not pairs:
        raise ValueError("没有可见列可导出")

    pairs = _dedupe_headers(pairs)
    orig_cols = [p[0] for p in pairs]
    headers = [p[1] for p in pairs]
    out = df_full.loc[:, orig_cols].copy()
    out.columns = headers
    return out


# ---- Excel ----


def dataframe_to_excel_bytes(df: pd.DataFrame) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "数据"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    body_align = Alignment(vertical="top", wrap_text=True)

    ncols = len(df.columns)
    for j, h in enumerate(df.columns, start=1):
        cell = ws.cell(1, j, str(h))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            cell = ws.cell(i, j, val)
            cell.alignment = body_align

    for j in range(1, ncols + 1):
        letter = get_column_letter(j)
        max_len = len(str(ws.cell(1, j).value))
        for i in range(2, ws.max_row + 1):
            v = ws.cell(i, j).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- PDF（reportlab）----


def _project_root() -> str:
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def _collect_existing_pdf_font_paths() -> List[str]:
    """按优先级列出可能存在的中文字体文件路径。"""
    base = _project_root()
    raw: List[str] = [
        os.path.join(base, PDF_FONT_USER_DIR, "SimHei.ttf"),
        os.path.join(base, PDF_FONT_USER_DIR, "simhei.ttf"),
        os.path.join(base, PDF_FONT_USER_DIR, "msyh.ttc"),
        os.path.join(base, PDF_FONT_USER_DIR, "MSYH.TTC"),
        os.path.join(base, "simhei.ttf"),
        os.path.join(base, "static", "fonts", "SimHei.ttf"),
    ]
    windir = os.environ.get("WINDIR") or os.path.join(
        os.environ.get("SystemRoot", r"C:\Windows")
    )
    win_fonts = os.path.join(windir, "Fonts")
    for name in (
        "simhei.ttf",
        "SIMHEI.TTF",
        "msyh.ttc",
        "MSYH.TTC",
        "msyhbd.ttc",
        "simsun.ttc",
        "SIMSUN.TTC",
        "simkai.ttf",
        "STXIHEI.TTF",
    ):
        raw.append(os.path.join(win_fonts, name))
    if sys.platform == "darwin":
        raw.extend(
            [
                "/System/Library/Fonts/STHeiti Medium.ttc",
                "/System/Library/Fonts/STHeiti Light.ttc",
                "/Library/Fonts/Microsoft/SimHei.ttf",
            ]
        )
    raw.extend(
        [
            "/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/noto/NotoSansCJK-Regular.ttc",
            "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        ]
    )
    seen: set = set()
    out: List[str] = []
    for p in raw:
        try:
            ap = os.path.normcase(os.path.abspath(p))
        except Exception:
            ap = p
        if os.path.isfile(p) and ap not in seen:
            seen.add(ap)
            out.append(p)
    return out


def _try_register_pdf_font(path: str) -> bool:
    """将 path 注册为 ReportLab 字体名 ExportSimHei；.ttc 尝试多个 subfontIndex。"""
    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except ImportError:
        return False

    if "ExportSimHei" in pdfmetrics.getRegisteredFontNames():
        return True

    lower = path.lower()
    if lower.endswith(".ttc"):
        for idx in (0, 1, 2, 3):
            try:
                pdfmetrics.registerFont(TTFont("ExportSimHei", path, subfontIndex=idx))
                logger.info("PDF 中文字体已注册: %s (subfontIndex=%s)", path, idx)
                return True
            except Exception:
                continue
        return False
    try:
        pdfmetrics.registerFont(TTFont("ExportSimHei", path))
        logger.info("PDF 中文字体已注册: %s", path)
        return True
    except Exception as ex:
        logger.debug("跳过字体 %s: %s", path, ex)
        return False


def register_pdf_chinese_font() -> Tuple[str, bool]:
    """
    注册 PDF 用中文字体。
    返回 (reportlab 字体名, 是否回退到 Helvetica)。
    若回退，logger.warning 并提示用户放置字体到 fonts/ 目录。
    """
    try:
        from reportlab.pdfbase import pdfmetrics
    except ImportError:
        logger.warning("未安装 reportlab，无法导出 PDF")
        return "Helvetica", True

    if "ExportSimHei" in pdfmetrics.getRegisteredFontNames():
        return "ExportSimHei", False

    for path in _collect_existing_pdf_font_paths():
        if _try_register_pdf_font(path):
            return "ExportSimHei", False

    logger.warning(PDF_FONT_HELP_ZH)
    return "Helvetica", True


def dataframe_to_pdf_bytes(df: pd.DataFrame) -> Tuple[bytes, bool]:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

    font_name, font_fallback = register_pdf_chinese_font()
    ncols = max(1, len(df.columns))
    pagesize = landscape(A4) if ncols > 10 else A4
    fs = 6 if ncols > 14 else 7 if ncols > 10 else 8

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=pagesize,
        leftMargin=22,
        rightMargin=22,
        topMargin=28,
        bottomMargin=28,
        title="数据导出",
    )

    headers = [str(c) for c in df.columns]
    data_rows: List[List[str]] = [headers]
    for _, row in df.iterrows():
        line = []
        for v in row:
            if v is None or (isinstance(v, float) and pd.isna(v)):
                line.append("")
            else:
                line.append(str(v))
        data_rows.append(line)

    col_widths = None
    if ncols:
        usable = pagesize[0] - doc.leftMargin - doc.rightMargin - 4
        per = usable / ncols
        col_widths = [max(per, 28)] * ncols

    tbl = Table(data_rows, repeatRows=1, colWidths=col_widths)
    tbl.setStyle(
        TableStyle(
            [
                ("FONT", (0, 0), (-1, -1), font_name, fs),
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("FONT", (0, 0), (-1, 0), font_name, fs + 1),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
            ]
        )
    )
    doc.build([tbl])
    return buf.getvalue(), font_fallback


# ---- CSV ----


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """UTF-8 BOM，便于 Excel 正确打开中文。"""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8")


# ---- PNG（matplotlib 表格）----


def dataframe_to_png_bytes(df: pd.DataFrame, max_rows: int = 400) -> bytes:
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    plt.rcParams["font.sans-serif"] = [
        "SimHei",
        "Microsoft YaHei",
        "Noto Sans CJK SC",
        "DejaVu Sans",
    ]

    view = df.head(max_rows)
    nrows, ncols = view.shape[0], view.shape[1]
    if ncols == 0:
        raise ValueError("没有列可绘制")

    cell_text = []
    for _, row in view.iterrows():
        cell_text.append(["" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v) for v in row])

    col_labels = [str(c) for c in view.columns]
    fig_w = min(56, max(10, ncols * 1.15))
    fig_h = min(40, max(4, 0.35 * (nrows + 2)))
    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")
    table = ax.table(
        cellText=cell_text,
        colLabels=col_labels,
        loc="center",
        cellLoc="center",
    )
    table.auto_set_font_size(False)
    table.set_fontsize(7)
    table.scale(1, 1.2)

    buf = io.BytesIO()
    fig.savefig(buf, format="png", bbox_inches="tight", dpi=120, pad_inches=0.2)
    plt.close(fig)
    buf.seek(0)
    return buf.getvalue()


def sanitize_export_basename(name: str) -> str:
    s = re.sub(r'[\x00-\x1f<>:"/\\|?*]+', "_", name).strip(" .") or "export"
    return s[:80]
